import sys
import os
import openpyxl
import glob
import unittest
import logging
from datetime import datetime

current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.abspath(os.path.join(current_dir, '..', '..'))
sys.path.append(project_root)

from dotenv import load_dotenv
dotenv_path = os.path.join(project_root, '.env')
load_dotenv(dotenv_path)

from src.utility.DownloadTheData import DownloadData

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class FixedTermDealTestCases(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        logger.info("Data is Downloading....")

        expected_data_path = os.path.join(current_dir, '..', 'tempStorage', 'expectedData')
        actual_data_path = os.path.join(current_dir, '..', 'tempStorage', 'actualData')

        cls.expected_data_dir = os.path.normpath(expected_data_path)
        cls.actual_data_dir = os.path.normpath(actual_data_path)

        downloader = DownloadData()
        downloader.get_files("DEV/Saving/Inbound/Flagstone Savings Spreadsheets", cls.expected_data_dir)
        downloader.get_file("SAVING_DEV.CONF.FLAGSTONE_FIXED_TERM_SAVING.xlsx", "DEV/Dev Data Out/Saving (Deposit)", cls.actual_data_dir)

        logger.info("Data Downloaded")

    @staticmethod
    def load_expected_excel(expectedExcel_file_path):
        try:
            logger.info(f"Loading expected Excel file:\n{expectedExcel_file_path}")
            expected_workbook = openpyxl.load_workbook(expectedExcel_file_path, data_only=True)
            if 'FTD' in expected_workbook.sheetnames:
                sheet = expected_workbook['FTD']
                return sheet
            else:
                logger.error(f"Workbook '{expectedExcel_file_path}' does not contain an 'FTD' sheet.")
                return None

        except Exception as e:
            logger.error(f"Error loading {expectedExcel_file_path}: {e}")
            return None

    @staticmethod
    def load_actual_excel(actualExcel_file_path):
        try:
            logger.info(f"Loading actual Excel file:\n{actualExcel_file_path}")
            actual_workbook = openpyxl.load_workbook(actualExcel_file_path, data_only=True)
            sheet = actual_workbook.active
            return sheet

        except Exception as e:
            logger.error(f"Error loading {actualExcel_file_path}: {e}")
            return None

    @classmethod
    def normalize_value(cls, value):
        # Normalize term length
        if isinstance(value, str) and value.endswith('M'):
            value = value.lstrip('0')  # Remove leading zeros

        # Normalize floating point precision
        try:
            value = round(float(value), 2)  # Round to 2 decimal places for comparison
        except ValueError:
            pass  # If not a number, keep the value as is

        # Normalize date/time formats
        try:
            if 'T' in value:  # For with "T"
                value = datetime.strptime(value, '%Y-%m-%dT%H:%M:%S.%fZ').strftime('%Y-%m-%d %H:%M:%S')
            else:
                value = datetime.strptime(value, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
        except (ValueError, TypeError):
            pass  # If not a date, keep the value as is

        return value

    @classmethod
    def get_row_values(cls, sheet, start_row, remove_first_column=False, remove_last_n_columns=0):
        rows = []
        for row in range(start_row, sheet.max_row + 1):
            row_values = []
            for col in range(1, sheet.max_column + 1):
                if remove_first_column and col == 1:
                    continue
                if remove_last_n_columns and col > sheet.max_column - remove_last_n_columns:
                    continue
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is not None:
                    normalized_value = cls.normalize_value(str(cell_value).strip())
                    row_values.append(normalized_value)
            # Only add rows that are not empty
            if any(row_values):
                rows.append(tuple(row_values))  # Store each row as a tuple for easier comparison
        return rows

    @staticmethod
    def compare_rows(row1, row2, tolerance=1e-6):
        if len(row1) != len(row2):
            return False
        for v1, v2 in zip(row1, row2):
            try:
                if abs(float(v1) - float(v2)) > tolerance:
                    return False
            except ValueError:
                if v1 != v2:
                    return False
        return True

    def test_to_validate_Flagstone_FTD_data_with_Kleene_output_data(self):
        all_expected_files = os.listdir(self.expected_data_dir)
        all_expected_rows = []

        # Dictionary to store the expected rows along with their source file
        expected_file_rows = {}

        # Read and aggregate data from all expected files
        for file_name in all_expected_files:
            expectedExcel_file_path = os.path.join(self.expected_data_dir, file_name)
            expected_sheet = self.load_expected_excel(expectedExcel_file_path)

            if expected_sheet is not None:
                expected_rows = self.get_row_values(expected_sheet, 4)
                all_expected_rows.extend(expected_rows)
                for row in expected_rows:
                    expected_file_rows[row] = file_name
            else:
                logger.error(f"Error loading expected sheet: {expectedExcel_file_path}")

        # Read actual data
        actualExcel_file_path = os.path.join(self.actual_data_dir, "SAVING_DEV.CONF.FLAGSTONE_FIXED_TERM_SAVING.xlsx")
        actual_sheet = self.load_actual_excel(actualExcel_file_path)

        if actual_sheet is not None:
            actual_rows = self.get_row_values(actual_sheet, 2, remove_first_column=True, remove_last_n_columns=4)

            # Validate if all_expected_rows are present in actual_rows
            matched_rows_count = 0
            missing_rows = []
            for expected_row in all_expected_rows:
                found = False
                for actual_row in actual_rows:
                    if self.compare_rows(expected_row, actual_row):
                        found = True
                        matched_rows_count += 1
                        break
                if not found:
                    missing_rows.append((expected_row, expected_file_rows[expected_row]))

            logger.info(f"Number of matched rows: {matched_rows_count}")

            # Log and report all missing rows
            if missing_rows:
                error_msgs = []
                for row, file_name in missing_rows:
                    error_msgs.append(f"Row '{row}' from file '{file_name}' not found in actual data.")

                error_msgs.append(f"Number of matched rows: {matched_rows_count}")

                # Join all error messages into a single message
                error_msg = "<br/>".join(error_msgs)
                logger.error(error_msg)

                # Fail the test with the error message
                self.fail(error_msg)

        else:
            logger.error("Error loading actual sheet.")
            self.fail("Error loading actual sheet.")

    @classmethod
    def tearDownClass(cls):
        logger.info("Tearing down resources")
        cls.clean_up_downloaded_files(cls.expected_data_dir)
        cls.clean_up_downloaded_files(cls.actual_data_dir)

    @classmethod
    def clean_up_downloaded_files(cls, folder_path):
        logger.info(f"Cleaning up files in {folder_path}...")
        files = glob.glob(os.path.join(folder_path, "*"))
        for file in files:
            os.remove(file)
            logger.info(f"Deleted file: {file}")
        logger.info("Clean up completed.")


if __name__ == '__main__':
    unittest.main()

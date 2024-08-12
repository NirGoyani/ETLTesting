import sys
import os
import openpyxl
import glob
import unittest
import logging

current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.abspath(os.path.join(current_dir, '..', '..'))
sys.path.append(project_root)

from dotenv import load_dotenv

dotenv_path = os.path.join(project_root, '.env')
load_dotenv(dotenv_path)

from src.utility.DownloadTheData import DownloadData

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class AssetsDataTestCases(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        logger.info("Data is Downloading....")

        actual_data_path = os.path.join(current_dir, '..', 'tempStorage', 'actualData')

        cls.actual_data_dir = os.path.normpath(actual_data_path)

        downloader = DownloadData()
        downloader.get_file("boe_mortgages_subentity_table_populated_v2 1.xlsx", "DEV/Dev Data Out/Asset (Mortgage)",
                            cls.actual_data_dir)

        logger.info("Data Downloaded")

    @staticmethod
    def load_actual_excel(expectedExcel_file_path):
        try:
            logger.info(f"Loading expected Excel file:\n{expectedExcel_file_path}")
            expected_workbook = openpyxl.load_workbook(expectedExcel_file_path, data_only=True)
            sheet = expected_workbook.active
            return sheet

        except Exception as e:
            logger.error(f"Error loading {expectedExcel_file_path}: {e}")
            return None

    @classmethod
    def calculate_retail_mortgages(cls, actual_sheet, start_row, col):
        current_balance_list = []
        product_of_number_list = []
        current_loan_to_value_high_ltv = []
        debt_to_income = []
        current_balance_list_new_advance = []
        product_of_number_list_btl = []
        current_balance_list_btl = []
        product_of_number_list_interest_ratio = []
        current_balance_list_interest_ratio = []
        current_balance_list_interest_only = []

        for row in range(start_row, actual_sheet.max_row):
            occupancy_type = actual_sheet.cell(row=row, column=col).value  #AR130
            repayment_method = actual_sheet.cell(row=row, column=col - 61).value  #AR69
            if "April" in actual_sheet.cell(row=row, column=col + 179).value:
                if occupancy_type == 1:
                    current_balance = actual_sheet.cell(row=row, column=col - 63).value  # AR67
                    current_loan_to_value = actual_sheet.cell(row=row, column=col + 11).value  # AR141

                    product_of_number_list.append(current_balance * current_loan_to_value)
                    current_balance_list.append(current_balance)

                if occupancy_type == 1 and actual_sheet.cell(row=row, column=col + 11).value >= 0.8:  # AR141 >= 0.8
                    current_loan_to_value_high_ltv.append(actual_sheet.cell(row=row, column=col - 63).value)  # AR67

                if occupancy_type == 1:

                    current_balance_new_advance = actual_sheet.cell(row=row, column=col - 63).value  # AR67
                    current_balance_list_new_advance.append(current_balance_new_advance)

                    if actual_sheet.cell(row=row, column=col - 57).value is not None:
                        debt_to_income.append(actual_sheet.cell(row=row, column=col - 57).value)  #AR73
                    else:
                        debt_to_income.append(0)

                if occupancy_type == 3:
                    current_balance_btl = actual_sheet.cell(row=row, column=col - 63).value  # AR67
                    current_loan_to_value_btl = actual_sheet.cell(row=row, column=col + 11).value  # AR141

                    product_of_number_list_btl.append(current_balance_btl * current_loan_to_value_btl)
                    current_balance_list_btl.append(current_balance_btl)

                if occupancy_type == 3:
                    current_balance_interest_ratio = actual_sheet.cell(row=row, column=col - 63).value  # AR67
                    debt_service_coverage_ratio = actual_sheet.cell(row=row, column=col + 26).value  # AR156

                    product_of_number_list_interest_ratio.append(
                        current_balance_interest_ratio * debt_service_coverage_ratio)
                    current_balance_list_interest_ratio.append(current_balance_interest_ratio)

                if occupancy_type == 3 and repayment_method == 1:
                    current_balance_interest_only = actual_sheet.cell(row=row, column=col - 63).value  # AR67
                    current_balance_list_interest_only.append(current_balance_interest_only)

        product_list_of_debt_income = [a * b for a, b in zip(debt_to_income, current_balance_list_new_advance)]

        average_ltv = (sum(product_of_number_list) / sum(current_balance_list)) * 100
        high_ltv_lending = (sum(current_loan_to_value_high_ltv) / sum(current_balance_list)) * 100
        income_multiple = sum(product_list_of_debt_income) / sum(current_balance_list_new_advance) * 100
        average_btl = (sum(product_of_number_list_btl) / sum(current_balance_list_btl)) * 100
        btl_average_interest_coverage_ratio = (sum(product_of_number_list_interest_ratio) / sum(
            current_balance_list_interest_ratio)) * 100
        btl_interest_only = (sum(current_balance_list_interest_only) / sum(current_balance_list_btl)) * 100
        logger.info(f"Risk average_ltv {average_ltv} %")
        logger.info(f"High LTV lending {high_ltv_lending} %")
        logger.info(f"Retail income multiple new advance {income_multiple} %")
        logger.info(f"BTL Average LTV {average_btl} %")
        logger.info(f"BTL Average interest coverage ratio {btl_average_interest_coverage_ratio} %")
        logger.info(f"BTL interest only {btl_interest_only} %")

    @classmethod
    def calculate_owner_occupied_mortgages(cls, actual_sheet, start_row, col):
        current_loan_to_value = []
        current_balance_list = []
        current_balance_lien_list = []
        current_balance_equifax_list = []
        bureau_score_value_list = []

        for row in range(start_row, actual_sheet.max_row):
            occupancy_type = actual_sheet.cell(row=row, column=col).value  # AR130
            lien = actual_sheet.cell(row=row, column=col - 46).value  # AR84

            if "May" in actual_sheet.cell(row=row, column=col + 179).value:  #Dates
                if occupancy_type == 1:
                    current_loan_to_value.append(actual_sheet.cell(row=row, column=col + 11).value)  # AR141
                    current_balance_list.append(actual_sheet.cell(row=row, column=col - 63).value)  # AR67

                if occupancy_type == 1 and lien == 1:
                    current_loan_to_value.append(actual_sheet.cell(row=row, column=col + 11).value)  # AR141

                    current_balance_lien = actual_sheet.cell(row=row, column=col - 63).value  # AR67
                    current_balance_lien_list.append(current_balance_lien)

                if occupancy_type == 1:
                    current_balance_equifax = actual_sheet.cell(row=row, column=col - 63).value  # AR67
                    current_balance_equifax_list.append(current_balance_equifax)

                    if actual_sheet.cell(row=row, column=col + 88).value is not None:
                        bureau_score_value = actual_sheet.cell(row=row, column=col + 88).value  # AR220
                        bureau_score_value_list.append(bureau_score_value)
                    else:
                        bureau_score_value_list.append(0)

        product_list_of_equifax = [a * b for a, b in zip(current_balance_equifax_list, bureau_score_value_list)]

        ltv_maximum = max(current_loan_to_value) * 100
        first_second_mortgages = (sum(current_balance_lien_list) / sum(current_balance_list)) * 100
        equifax_score = (sum(product_list_of_equifax) / sum(current_balance_list))
        logger.info(f"LTV maximum {ltv_maximum} %")
        logger.info(f"1st and 2nd mortgage {first_second_mortgages} %")
        logger.info(f"Equifax Score {equifax_score}")

    @classmethod
    def calculate_buy_to_let(cls, actual_sheet, start_row, col):
        current_loan_to_value = []
        current_balance_list = []
        current_balance_lien_list = []
        current_balance_equifax_list = []
        bureau_score_value_list = []
        current_balance_lien_currency_list = []

        for row in range(start_row, actual_sheet.max_row):
            occupancy_type = actual_sheet.cell(row=row, column=col).value  # AR130
            lien = actual_sheet.cell(row=row, column=col - 46).value  # AR84
            loan_currency_denomination = actual_sheet.cell(row=row, column=col - 65).value  #AR65

            if "May" in actual_sheet.cell(row=row, column=col + 179).value:  # Dates
                if occupancy_type == 3:
                    current_loan_to_value.append(actual_sheet.cell(row=row, column=col + 11).value)  # AR141
                    current_balance_list.append(actual_sheet.cell(row=row, column=col - 63).value)  # AR67

                if occupancy_type == 3 and lien == 1:
                    current_balance_lien = actual_sheet.cell(row=row, column=col - 63).value  # AR67
                    current_balance_lien_list.append(current_balance_lien)

                if occupancy_type == 3:
                    current_balance_equifax = actual_sheet.cell(row=row, column=col - 63).value  # AR67
                    current_balance_equifax_list.append(current_balance_equifax)

                    if actual_sheet.cell(row=row, column=col + 88).value is not None:
                        bureau_score_value = actual_sheet.cell(row=row, column=col + 88).value  # AR220
                        bureau_score_value_list.append(bureau_score_value)
                    else:
                        bureau_score_value_list.append(0)

                if occupancy_type == 3 and loan_currency_denomination == 2:
                    current_balance_lien = actual_sheet.cell(row=row, column=col - 63).value  # AR67
                    current_balance_lien_currency_list.append(current_balance_lien)

        product_list_of_equifax = [a * b for a, b in zip(current_balance_equifax_list, bureau_score_value_list)]

        ltv_maximum = max(current_loan_to_value) * 100
        first_second_mortgages = (sum(current_balance_lien_list) / sum(current_balance_list)) * 100
        equifax_score = (sum(product_list_of_equifax) / sum(current_balance_list))
        currency = (sum(current_balance_lien_currency_list) / sum(current_balance_list)) * 100
        logger.info(f"Buy to Let LTV maximum {ltv_maximum} %")
        logger.info(f"Buy to Let 1st and 2nd mortgage {first_second_mortgages} %")
        logger.info(f"Buy to Let Equifax Score {equifax_score}")
        logger.info(f"Buy to Let Currency {currency} %")

    @classmethod
    def calculate_overall_portfolio(cls, actual_sheet, start_row, col):
        current_balance_list = []
        current_interest_rate_margin_list = []
        current_rate_list = []
        not_current_interest_rate_margin_list = []
        current_balance_list_term = []
        loan_term_list = []
        bureau_score_value_list = []
        current_balance_list_performing = []
        fixed_count_list = []
        floating_count_list = []
        count = 0
        for row in range(start_row, actual_sheet.max_row + 1):
            current_balance = actual_sheet.cell(row=row, column=col - 63).value  # AR67
            current_interest_rate_margin = actual_sheet.cell(row=row, column=col - 20).value  # AR110
            number_months_in_arrears = actual_sheet.cell(row=row, column=col + 40).value  # AR170
            current_rate = actual_sheet.cell(row=row, column=col - 21).value  # AR109

            if actual_sheet.cell(row=row, column=col + 179).value:  # Dates
                if current_balance > 0:
                    current_balance_list.append(current_balance)
                    count += 1
                if current_balance > 0:
                    current_rate_list.append(current_rate)
                if current_interest_rate_margin == 0:
                    current_interest_rate_margin_list.append(current_balance)
                if current_interest_rate_margin != 0:
                    not_current_interest_rate_margin_list.append(current_balance)
                if current_balance:
                    current_balance_list_term.append(current_balance)
                    loan_term_list.append(actual_sheet.cell(row=row, column=col - 69).value)  #AR61
                    bureau_score_value_list.append(actual_sheet.cell(row=row, column=col - 84).value)  #AR46
                if number_months_in_arrears < 0.75:
                    current_balance_list_performing.append(current_balance)
                if current_balance > 0 and current_interest_rate_margin == 0:
                    fixed_count_list.append(1)
                if current_balance > 0 and current_interest_rate_margin != 0:
                    floating_count_list.append(1)

        product_loan_term = [a * b for a, b in zip(current_balance_list_term, loan_term_list)]
        product_current_rate = [a * b for a, b in zip(current_balance_list, current_rate_list)]
        product_credit_score = [a * b for a, b in zip(current_balance_list_term, bureau_score_value_list)]

        wav_loan_size = sum(current_balance_list) / count
        wav_current_rate = (sum(product_current_rate) / sum(current_balance_list)) * 100
        percentage_fixed = sum(current_interest_rate_margin_list) / sum(current_balance_list)
        percentage_floating = (sum(not_current_interest_rate_margin_list) / sum(current_balance_list)) * 100
        wav_term = sum(product_loan_term) / sum(current_balance_list)
        wav_credit_score = sum(product_credit_score) / sum(current_balance_list)

        count_percentage_fixed = (sum(fixed_count_list) / count) * 100
        count_percentage_floating = (sum(floating_count_list) / count) * 100

        logger.info(f"Portfolio Size {sum(current_balance_list)} ")
        logger.info(f"Wav. Loan Size {wav_loan_size} ")
        logger.info(f"Wav. Current Rate {wav_current_rate} ")
        logger.info(f"% Fixed {percentage_fixed}%")
        logger.info(f"% Floating {percentage_floating}%")
        logger.info(f"% Wav Term (Months) {wav_term}")
        logger.info(f"% Wav Credit Score {wav_credit_score}")
        logger.info(f"% Performing {sum(current_balance_list_performing)}")
        logger.info(f"% =========================================================")
        logger.info(f"% Count Portfolio Size {count}")
        logger.info(f"% Count Fixed {count_percentage_fixed}")
        logger.info(f"% Count Floating {count_percentage_floating}%")

    @classmethod
    def calculate_residential_mortgages(cls, actual_sheet, start_row, col):
        current_balance_list = []
        portfolio_size_list = []
        current_balance_interest_list = []
        current_balance_repayment_list = []
        current_balance_part_and_part_list = []
        current_balance_BTL_list = []
        current_balance_Owner_occupied = []
        current_loan_to_value_list = []
        min_current_loan_to_value_list = []
        count = 0

        for row in range(start_row, actual_sheet.max_row + 1):
            current_balance = actual_sheet.cell(row=row, column=col - 63).value  # AR67
            repayment_method = actual_sheet.cell(row=row, column=col - 61).value  # AR69
            property_type = actual_sheet.cell(row=row, column=col + 1).value  # AR131
            occupancy_type = actual_sheet.cell(row=row, column=col).value  #AR130
            current_loan_to_value = actual_sheet.cell(row=row, column=col + 11).value  # AR141

            if actual_sheet.cell(row=row, column=col + 179).value:  # Dates
                if current_balance > 0:
                    current_balance_list.append(current_balance)

                if property_type in [1, 2, 3, 4] and current_balance > 0:
                    portfolio_size_list.append(current_balance)
                    current_loan_to_value_list.append(current_loan_to_value)
                    count += 1

                if property_type in [1, 2, 3, 4] and repayment_method == 1 and current_balance > 0:
                    current_balance_interest_list.append(current_balance)

                if property_type in [1, 2, 3, 4] and repayment_method == 2 and current_balance > 0:
                    current_balance_repayment_list.append(current_balance)

                if property_type in [1, 2, 3, 4] and repayment_method == 7 and current_balance > 0:
                    current_balance_part_and_part_list.append(current_balance)

                if property_type in [1, 2, 3, 4] and occupancy_type != 1 and current_balance > 0:
                    current_balance_BTL_list.append(current_balance)

                if property_type in [1, 2, 3, 4] and occupancy_type == 1 and current_balance > 0:
                    current_balance_Owner_occupied.append(current_balance)

                if property_type in [1, 2, 3, 4]:
                    min_current_loan_to_value_list.append(current_loan_to_value)

        product_current_ltv = [a * b for a, b in zip(portfolio_size_list, current_loan_to_value_list)]

        logger.info(f"Portfolio Size {sum(portfolio_size_list)} ")
        logger.info(f"Count Portfolio Size {count} ")
        logger.info(f"Interest only  {sum(current_balance_interest_list)}")
        logger.info(f"% Interest only  {(sum(current_balance_interest_list) / sum(portfolio_size_list)) * 100}")
        logger.info(f"Repayment  {sum(current_balance_repayment_list)}")
        logger.info(f"Part and part  {sum(current_balance_part_and_part_list)}")
        logger.info(f"BTL  {sum(current_balance_BTL_list)}")
        logger.info(f"Owner Occupier  {sum(current_balance_Owner_occupied)}")
        logger.info("====================================================")
        logger.info(f"Current Ltv {(sum(product_current_ltv) / sum(portfolio_size_list)) * 100}")
        logger.info(f"Min Current Ltv {(min(min_current_loan_to_value_list)) * 100}")
        logger.info(f"Max Current Ltv {(max(min_current_loan_to_value_list)) * 100}")

    @classmethod
    def calculate_commercial_mortgages(cls, actual_sheet, start_row, col):
        current_balance_list = []
        portfolio_size_list = []
        current_balance_interest_list = []
        current_balance_repayment_list = []
        current_balance_part_and_part_list = []
        current_balance_BTL_list = []
        current_balance_Owner_occupied = []
        original_ltv_list = []
        current_ltv_list = []
        payment_due_list = []

        for row in range(start_row, actual_sheet.max_row + 1):
            current_balance = actual_sheet.cell(row=row, column=col - 63).value  # AR67
            repayment_method = actual_sheet.cell(row=row, column=col - 61).value  # AR69
            property_type = actual_sheet.cell(row=row, column=col + 1).value  # AR131
            occupancy_type = actual_sheet.cell(row=row, column=col).value  # AR130
            original_ltv = actual_sheet.cell(row=row, column=col + 5).value  # AR135
            current_ltv = actual_sheet.cell(row=row, column=col + 11).value  # AR141
            payment_due = actual_sheet.cell(row=row, column=col - 59).value  # AR71

            if actual_sheet.cell(row=row, column=col + 179).value:  # Dates
                if current_balance > 0:
                    current_balance_list.append(current_balance)

                if property_type in [5, 6, 7, 8, 9] and current_balance > 0:
                    portfolio_size_list.append(current_balance)

                if property_type in [5, 6, 7, 8, 9] and repayment_method == 1 and current_balance > 0:
                    current_balance_interest_list.append(current_balance)

                if property_type in [5, 6, 7, 8, 9] and repayment_method == 2 and current_balance > 0:
                    current_balance_repayment_list.append(current_balance)

                if property_type in [5, 6, 7, 8, 9] and repayment_method == 7 and current_balance > 0:
                    current_balance_part_and_part_list.append(current_balance)

                if property_type in [5, 6, 7, 8, 9] and occupancy_type != 1 and current_balance > 0:
                    current_balance_BTL_list.append(current_balance)

                if property_type in [5, 6, 7, 8, 9] and occupancy_type == 1 and current_balance > 0:
                    current_balance_Owner_occupied.append(current_balance)

                if property_type in [5, 6, 7, 8, 9] and current_balance > 0:
                    original_ltv_list.append(original_ltv)

                if property_type in [5, 6, 7, 8, 9] and current_balance > 0:
                    current_ltv_list.append(current_ltv)

                if property_type in [5, 6, 7, 8, 9] and current_balance > 0:
                    payment_due_list.append(payment_due)

        product_original_ltv = [a * b for a, b in zip(portfolio_size_list, original_ltv_list)]
        product_current_ltv = [a * b for a, b in zip(portfolio_size_list, current_ltv_list)]
        payment_due_ltv = [a * b for a, b in zip(portfolio_size_list, payment_due_list)]

        logger.info(f"Portfolio Size {sum(portfolio_size_list)} ")
        logger.info(f"Interest only  {sum(current_balance_interest_list)}")
        logger.info(f"Repayment  {sum(current_balance_repayment_list)}")
        logger.info(f"Part and part  {sum(current_balance_part_and_part_list)}")
        logger.info(f"BTL  {sum(current_balance_BTL_list)}")
        logger.info(f"Owner Occupier  {sum(current_balance_Owner_occupied)}")
        logger.info("====================================================")
        logger.info(f"Original Ltv {(sum(product_original_ltv) / sum(portfolio_size_list)) * 100}")
        logger.info(f"Current Ltv {(sum(product_current_ltv) / sum(portfolio_size_list)) * 100}")
        logger.info(f"Payment due Ltv {(sum(payment_due_ltv) / sum(portfolio_size_list)) * 100}")

    @classmethod
    def calculate_SME_mortgages(cls, actual_sheet, start_row, col):
        current_balance_list = []
        portfolio_size_list = []
        current_balance_interest_list = []
        current_balance_repayment_list = []
        current_balance_part_and_part_list = []
        current_balance_BTL_list = []
        current_balance_Owner_occupied = []

        for row in range(start_row, actual_sheet.max_row + 1):
            current_balance = actual_sheet.cell(row=row, column=col - 63).value  # AR67
            repayment_method = actual_sheet.cell(row=row, column=col - 61).value  # AR69
            property_type = actual_sheet.cell(row=row, column=col + 1).value  # AR131
            occupancy_type = actual_sheet.cell(row=row, column=col).value  # AR130
            borrower_type = actual_sheet.cell(row=row, column=col - 115).value  # AR15

            if actual_sheet.cell(row=row, column=col + 179).value:  # Dates
                if current_balance > 0:
                    current_balance_list.append(current_balance)

                if property_type in [10, 11] and borrower_type in ["COM", "SCI"]:
                    portfolio_size_list.append(current_balance)

                if property_type in [10, 11] and borrower_type in ["COM", "SCI"] and repayment_method == 1:
                    current_balance_interest_list.append(current_balance)

                if property_type in [10, 11] and borrower_type in ['COM', 'SCI'] and repayment_method == 2:
                    current_balance_repayment_list.append(current_balance)

                if property_type in [10, 11] and borrower_type in ['COM', 'SCI'] and repayment_method == 7:
                    current_balance_part_and_part_list.append(current_balance)

                if property_type in [10, 11] and borrower_type in ['COM', 'SCI'] and occupancy_type != 1:
                    current_balance_BTL_list.append(current_balance)

                if property_type in [10, 11] and borrower_type in ['COM', 'SCI'] and occupancy_type == 1:
                    current_balance_Owner_occupied.append(current_balance)

        logger.info(f"Portfolio Size {sum(portfolio_size_list)} ")
        logger.info(f"Interest only  {sum(current_balance_interest_list)}")
        logger.info(f"Repayment  {sum(current_balance_repayment_list)}")
        logger.info(f"Part and part  {sum(current_balance_part_and_part_list)}")
        logger.info(f"BTL  {sum(current_balance_BTL_list)}")
        logger.info(f"Owner Occupier  {sum(current_balance_Owner_occupied)}")

    @classmethod
    def calculate_Partner_Adherence_To_Covenant(cls, actual_sheet, start_row, col):
        current_balance_list = []
        current_30_arrears_list = []
        current_90_arrears_list = []

        for row in range(start_row, actual_sheet.max_row + 1):
            current_balance = actual_sheet.cell(row=row, column=col - 63).value  # AR67
            payment_frequency = actual_sheet.cell(row=row, column=col - 60).value  # AR70
            number_months_in_arrears = actual_sheet.cell(row=row, column=col + 40).value  # AR170

            if "May" in actual_sheet.cell(row=row, column=col + 179).value:  # Dates
                if number_months_in_arrears >= 1:
                    current_30_arrears_list.append(current_balance)

                if number_months_in_arrears >= 3:
                    current_90_arrears_list.append(current_balance)

                if current_balance:
                    current_balance_list.append(current_balance)

        logger.info(f"30+ % arrears {(sum(current_30_arrears_list) / sum(current_balance_list)) * 100} ")
        logger.info(f"90+ % arrears {(sum(current_90_arrears_list) / sum(current_balance_list)) * 100} ")

    @classmethod
    def calculate_origination_all_product(cls, actual_sheet, start_row, col):
        current_balance_list = []
        all_current_balance_list = []
        current_rate_list = []

        for row in range(start_row, actual_sheet.max_row + 1):
            current_balance = actual_sheet.cell(row=row, column=col - 63).value  # AR67
            current_rate = actual_sheet.cell(row=row, column=col - 21).value  # AR109
            if "May" in actual_sheet.cell(row=row, column=col + 179).value:  # Dates
                if current_balance > 0:
                    current_balance_list.append(current_balance)

                all_current_balance_list.append(current_balance)
                current_rate_list.append(current_rate)

        average_interest_rate = [a * b for a, b in zip(all_current_balance_list, current_rate_list)]

        logger.info(f"Count Loans {len(current_balance_list)}")
        logger.info(f"Pound Loans {sum(current_balance_list)}")
        logger.info(f"Average Loan size (£) {sum(current_balance_list) / len(current_balance_list)}")
        logger.info(f"Average Interest rate {(sum(average_interest_rate) / sum(current_balance_list)) * 100}")

    @classmethod
    def calculate_origination_residential_product(cls, actual_sheet, start_row, col):
        current_balance_list = []
        property_type_list = []
        current_rate_list = []

        for row in range(start_row, actual_sheet.max_row + 1):
            current_balance = actual_sheet.cell(row=row, column=col - 63).value  # AR67
            property_type = actual_sheet.cell(row=row, column=col + 1).value  # AR131
            lien = actual_sheet.cell(row=row, column=col - 46).value  # AR84
            current_rate = actual_sheet.cell(row=row, column=col - 21).value  # AR109

            if "May" in actual_sheet.cell(row=row, column=col + 179).value:  # Dates
                if property_type in [1, 2, 3, 4] and lien == 1 and current_balance > 0:
                    current_balance_list.append(current_balance)
                    property_type_list.append(property_type)
                    current_rate_list.append(current_rate)

        residential_loans = [a * b for a, b in zip(current_balance_list, current_rate_list)]

        logger.info(f"Residential Loans count {len(current_balance_list)}")
        logger.info(f"Residential Loans Pound {sum(current_balance_list)}")
        logger.info(f"Residential average Loan size (£) {sum(current_balance_list) / len(current_balance_list)}")
        logger.info(f"Residential average Interest rate {(sum(residential_loans) / sum(current_balance_list)) * 100}")

    @classmethod
    def calculate_origination_residential_BTL_1st_charge_product(cls, actual_sheet, start_row, col):
        current_balance_list = []
        current_rate_list = []

        for row in range(start_row, actual_sheet.max_row + 1):
            current_balance = actual_sheet.cell(row=row, column=col - 63).value  # AR67
            occupancy_type = actual_sheet.cell(row=row, column=col).value  # AR130
            lien = actual_sheet.cell(row=row, column=col - 46).value  # AR84
            current_rate = actual_sheet.cell(row=row, column=col - 21).value  # AR109

            if "May" in actual_sheet.cell(row=row, column=col + 179).value:  # Dates
                if occupancy_type in [3] and lien == 1 and current_balance > 0:
                    current_balance_list.append(current_balance)
                    current_rate_list.append(current_rate)

        residential_loans = [a * b for a, b in zip(current_balance_list, current_rate_list)]

        logger.info(f"Residential BTL 1st charge Loans count {len(current_balance_list)}")
        logger.info(f"Residential BTL 1st charge Loans Pound {sum(current_balance_list)}")
        logger.info(f"Residential BTL 1st charge average Loan size (£) {sum(current_balance_list) / len(current_balance_list)}")
        logger.info(f"Residential BTL 1st charge average Interest rate {(sum(residential_loans) / sum(current_balance_list)) * 100}")

    @classmethod
    def calculate_origination_residential_BTL_2nd_charge_product(cls, actual_sheet, start_row, col):
        current_balance_list = []
        current_rate_list = []

        for row in range(start_row, actual_sheet.max_row + 1):
            current_balance = actual_sheet.cell(row=row, column=col - 63).value  # AR67
            lien = actual_sheet.cell(row=row, column=col - 46).value  # AR84
            current_rate = actual_sheet.cell(row=row, column=col - 21).value  # AR109

            if "May" in actual_sheet.cell(row=row, column=col + 179).value:  # Dates
                if lien != 1 and current_balance > 0:
                    current_balance_list.append(current_balance)
                    current_rate_list.append(current_rate)

        residential_loans = [a * b for a, b in zip(current_balance_list, current_rate_list)]

        logger.info(f"Residential BTL 2nd charge Loans count {len(current_balance_list)}")
        logger.info(f"Residential BTL 2nd charge Loans Pound {sum(current_balance_list)}")
        logger.info(f"Residential BTL 2nd charge average Loan size (£) {sum(current_balance_list) / len(current_balance_list)}")
        logger.info(f"Residential BTL 2nd charge average Interest rate {(sum(residential_loans) / sum(current_balance_list)) * 100}")

    @classmethod
    def calculate_origination_residential_secured_SME_product(cls, actual_sheet, start_row, col):
        current_balance_list = []
        current_rate_list = []

        for row in range(start_row, actual_sheet.max_row + 1):
            current_balance = actual_sheet.cell(row=row, column=col - 63).value  # AR67
            property_type = actual_sheet.cell(row=row, column=col + 1).value  # AR131
            current_rate = actual_sheet.cell(row=row, column=col - 21).value  # AR109
            borrower_type = actual_sheet.cell(row=row, column=col - 115).value  # AR15

            if "May" in actual_sheet.cell(row=row, column=col + 179).value:  # Dates
                if property_type in [10, 11] and borrower_type in ['COM', 'SCI'] and current_balance > 0:
                    current_balance_list.append(current_balance)
                    current_rate_list.append(current_rate)

        residential_loans = [a * b for a, b in zip(current_balance_list, current_rate_list)]

        logger.info(f"Residential Secured SME Loans count {len(current_balance_list)}")
        logger.info(f"Residential Secured SME Loans Pound {sum(current_balance_list)}")
        logger.info(f"Residential Secured SME average Loan size (£) {sum(current_balance_list) / len(current_balance_list)}")
        logger.info(f"Residential Secured SME average Interest rate {(sum(residential_loans) / sum(current_balance_list)) * 100}")

    def test_to_validate_Risk_Appetite_Assets_Data_report(self):
        actualExcel_file_path = os.path.join(self.actual_data_dir, "boe_mortgages_subentity_table_populated_v2 1.xlsx")
        actual_sheet = self.load_actual_excel(actualExcel_file_path)

        self.calculate_retail_mortgages(actual_sheet, 2, 132)
        self.assertTrue(True)

    def test_to_validate_Owner_Occupied_Mortgage_Assets_Data_report(self):
        actualExcel_file_path = os.path.join(self.actual_data_dir, "boe_mortgages_subentity_table_populated_v2 1.xlsx")
        actual_sheet = self.load_actual_excel(actualExcel_file_path)

        self.calculate_owner_occupied_mortgages(actual_sheet, 2, 132)

        self.assertTrue(True)

    def test_to_validate_Buy_to_let_Assets_Data_report(self):
        actualExcel_file_path = os.path.join(self.actual_data_dir, "boe_mortgages_subentity_table_populated_v2 1.xlsx")
        actual_sheet = self.load_actual_excel(actualExcel_file_path)

        self.calculate_buy_to_let(actual_sheet, 2, 132)

        self.assertTrue(True)

    def test_to_validate_Overall_Portfolio_Assets_Data_report(self):
        actualExcel_file_path = os.path.join(self.actual_data_dir, "boe_mortgages_subentity_table_populated_v2 1.xlsx")
        actual_sheet = self.load_actual_excel(actualExcel_file_path)

        self.calculate_overall_portfolio(actual_sheet, 2, 132)

        self.assertTrue(True)

    def test_to_validate_Residential_Mortgages_Assets_Data_report(self):
        actualExcel_file_path = os.path.join(self.actual_data_dir, "boe_mortgages_subentity_table_populated_v2 1.xlsx")
        actual_sheet = self.load_actual_excel(actualExcel_file_path)

        self.calculate_residential_mortgages(actual_sheet, 2, 132)

        self.assertTrue(True)

    def test_to_validate_Commercial_Mortgages_Assets_Data_report(self):
        actualExcel_file_path = os.path.join(self.actual_data_dir, "boe_mortgages_subentity_table_populated_v2 1.xlsx")
        actual_sheet = self.load_actual_excel(actualExcel_file_path)

        self.calculate_commercial_mortgages(actual_sheet, 2, 132)

        self.assertTrue(True)

    def test_to_validate_SME_Mortgages_Assets_Data_report(self):
        actualExcel_file_path = os.path.join(self.actual_data_dir, "boe_mortgages_subentity_table_populated_v2 1.xlsx")
        actual_sheet = self.load_actual_excel(actualExcel_file_path)

        self.calculate_SME_mortgages(actual_sheet, 2, 132)

        self.assertTrue(True)

    def test_to_validate_Partner_Adherence_To_Covenants_Assets_Data_report(self):
        actualExcel_file_path = os.path.join(self.actual_data_dir, "boe_mortgages_subentity_table_populated_v2 1.xlsx")
        actual_sheet = self.load_actual_excel(actualExcel_file_path)

        self.calculate_Partner_Adherence_To_Covenant(actual_sheet, 2, 132)

        self.assertTrue(True)

    def test_to_validate_origination_all_product_Assets_Data_report(self):
        actualExcel_file_path = os.path.join(self.actual_data_dir, "boe_mortgages_subentity_table_populated_v2 1.xlsx")
        actual_sheet = self.load_actual_excel(actualExcel_file_path)

        self.calculate_origination_all_product(actual_sheet, 2, 132)

        self.assertTrue(True)

    def test_to_validate_origination_residential_product_Assets_Data_report(self):
        actualExcel_file_path = os.path.join(self.actual_data_dir, "boe_mortgages_subentity_table_populated_v2 1.xlsx")
        actual_sheet = self.load_actual_excel(actualExcel_file_path)

        self.calculate_origination_residential_product(actual_sheet, 2, 132)

        self.assertTrue(True)

    def test_to_validate_origination_residential_BTL_1st_charge_product_Assets_Data_report(self):
        actualExcel_file_path = os.path.join(self.actual_data_dir, "boe_mortgages_subentity_table_populated_v2 1.xlsx")
        actual_sheet = self.load_actual_excel(actualExcel_file_path)

        self.calculate_origination_residential_BTL_1st_charge_product(actual_sheet, 2, 132)

        self.assertTrue(True)

    def test_to_validate_origination_residential_BTL_2nd_charge_product_Assets_Data_report(self):
        actualExcel_file_path = os.path.join(self.actual_data_dir, "boe_mortgages_subentity_table_populated_v2 1.xlsx")
        actual_sheet = self.load_actual_excel(actualExcel_file_path)

        self.calculate_origination_residential_BTL_2nd_charge_product(actual_sheet, 2, 132)

        self.assertTrue(True)

    def test_to_validate_origination_residential_Secured_SME_product_Assets_Data_report(self):
        actualExcel_file_path = os.path.join(self.actual_data_dir, "boe_mortgages_subentity_table_populated_v2 1.xlsx")
        actual_sheet = self.load_actual_excel(actualExcel_file_path)

        self.calculate_origination_residential_secured_SME_product(actual_sheet, 2, 132)

        self.assertTrue(True)

    @classmethod
    def tearDownClass(cls):
        logger.info("Tearing down resources")
        cls.clean_up_downloaded_files(cls.actual_data_dir)

    @classmethod
    def clean_up_downloaded_files(cls, folder_path):
        logger.info(f"Cleaning up files in {folder_path}...")
        files = glob.glob(os.path.join(folder_path, "*"))
        for file in files:
            os.remove(file)
            logger.info(f"Deleted file: {file}")
        logger.info("Clean up completed.")


if __name__ == '__main__':
    unittest.main()

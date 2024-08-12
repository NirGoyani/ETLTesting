import sys
import os
import glob
import unittest
import logging
import openpyxl
from datetime import datetime, timedelta

current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.abspath(os.path.join(current_dir, '..', '..'))
sys.path.append(project_root)

from dotenv import load_dotenv

dotenv_path = os.path.join(project_root, '.env')
load_dotenv(dotenv_path)

from src.utility.DownloadTheData import DownloadData

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class DailyInterestRateForFTDTestCases(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        logger.info("Data is Downloading....")

        expected_data_path = os.path.join(current_dir, '..', 'tempStorage', 'expectedData')
        actual_data_path = os.path.join(current_dir, '..', 'tempStorage', 'actualData')

        cls.expected_data_dir = os.path.normpath(expected_data_path)
        cls.actual_data_dir = os.path.normpath(actual_data_path)

        # To store all the error log
        cls.error_msgs = []
        cls.total_rows = 0
        cls.matched_rows = False

        downloader = DownloadData()
        downloader.get_files("DEV/Saving/Inbound/Flagstone Savings Spreadsheets", cls.expected_data_dir)
        downloader.get_file("SAVING_DEV.CONF.DAILY_INTEREST_RATE_FIXED_TERM.xlsx", "DEV/Dev Data Out/Saving (Deposit)", cls.actual_data_dir)

        logger.info("Data Downloaded")

    @staticmethod
    def load_expected_excel(expectedExcel_file_path):
        try:
            logger.info(f"Loading expected Excel file:\n{expectedExcel_file_path}")
            expected_workbook = openpyxl.load_workbook(expectedExcel_file_path, data_only=True)
            if 'FTD' in expected_workbook.sheetnames:
                sheet = expected_workbook['FTD']
                return sheet
            else:
                logger.error(f"Workbook '{expectedExcel_file_path}' does not contain an 'FTD' sheet.")
                return None

        except Exception as e:
            logger.error(f"Error loading {expectedExcel_file_path}: {e}")
            return None

    @staticmethod
    def load_actual_excel(actualExcel_file_path):
        try:
            logger.info(f"Loading actual Excel file:\n{actualExcel_file_path}")
            actual_workbook = openpyxl.load_workbook(actualExcel_file_path, data_only=True)
            sheet = actual_workbook.active
            return sheet

        except Exception as e:
            logger.error(f"Error loading {actualExcel_file_path}: {e}")
            return None

    @classmethod
    def get_row_values(cls, sheet, row, col):
        cell_value = sheet.cell(row=row, column=col).value
        return cell_value

    @classmethod
    def check_the_date(cls, expected_sheet, exp_row, exp_col, actual_sheet, actual_row, actual_col, no_of_days):
        expected_date = expected_sheet.cell(row=exp_row, column=exp_col).value

        for row in range(actual_row, actual_row + no_of_days):  # for 12M There are 365 days and for 01M there are 30 days
            actual_date = actual_sheet.cell(row=row, column=actual_col).value
            if expected_date == actual_date:
                expected_date += timedelta(days=1)
                cls.matched_rows = True
            else:
                logger.error(f"Date didn't matched {row}: expected {expected_date}, got {actual_date}")

    @classmethod
    def check_the_day_sequence(cls, actual_sheet, actual_row, actual_col, no_of_days):
        for day_sequence in range(1, no_of_days):
            cls.total_rows += 1
            actual_day = actual_sheet.cell(row=actual_row, column=actual_col).value
            if actual_day == day_sequence:
                actual_row += 1
                cls.matched_rows = True
            else:
                logger.error(f"Day sequence didn't matched")
                cls.error_msgs.append(f"Day sequence didn't matched")

    @classmethod
    def check_the_credit(cls, expected_sheet, exp_row, exp_col, actual_sheet, actual_row,actual_col):
        expected_credit = expected_sheet.cell(row=exp_row, column=exp_col).value
        actual_credit = actual_sheet.cell(row=actual_row, column=actual_col).value
        if expected_credit != actual_credit:
            logger.error(f"Credit amount didn't matched {actual_row}: expected {expected_credit}, got {actual_credit}")
            cls.error_msgs.append(f"Opening Balance pre interest didn't matched {actual_row}: expected {expected_credit}, got {actual_credit}")
            cls.matched_rows -= 1
        else:
           cls.matched_rows = True

    @classmethod
    def check_the_opening_balance_pre_interest(cls, expected_sheet,  exp_row, exp_col, actual_sheet, actual_row, actual_col, no_of_days):
        expected_opening_balance = 0

        for row in range(actual_row, actual_row + no_of_days-1):  # First row of opening balance is 0, so started from 2nd row
            credit = float(actual_sheet.cell(row=row, column=actual_col + 1).value)
            debit = float(actual_sheet.cell(row=row, column=actual_col + 2).value)
            opening_balance = credit - debit
            expected_opening_balance = expected_opening_balance + opening_balance
            actual_opening_balance = actual_sheet.cell(row=row+1, column=actual_col).value
            if round(expected_opening_balance,2) != actual_opening_balance:
                logger.error(f"Opening Balance pre interest didn't matched {row}: expected {expected_opening_balance}, got {actual_opening_balance}")
                cls.error_msgs.append(f"Opening Balance pre interest didn't matched {row}: expected {expected_opening_balance}, got {actual_opening_balance}")
            else:
               cls.matched_rows = True

    @classmethod
    def check_the_closing_balance_pre_interest(cls, expected_sheet,  exp_row, exp_col, actual_sheet, actual_row, actual_col, no_of_days):
        expected_closing_balance = 0

        for row in range(actual_row, actual_row + no_of_days):
            actual_closing_balance = actual_sheet.cell(row=row, column=actual_col).value
            credit = float(actual_sheet.cell(row=row, column=actual_col - 2).value)
            debit = float(actual_sheet.cell(row=row, column=actual_col - 1).value)
            closing_balance = credit - debit
            expected_closing_balance = expected_closing_balance + closing_balance
            if round(expected_closing_balance,2) != actual_closing_balance:
                logger.error(f"Closing Balance pre interest didn't matched at row {row}: expected {expected_closing_balance}, got {actual_closing_balance}")
                cls.error_msgs.append(f"Closing Balance pre interest didn't matched at row {row}: expected {expected_closing_balance}, got {actual_closing_balance}")
            else:
               cls.matched_rows = True

    @classmethod
    def check_the_interest_rate(cls, expected_sheet, exp_row, exp_col, actual_sheet, actual_row, actual_col, no_of_days):
        expected_interest_rate = round(expected_sheet.cell(row=exp_row, column=exp_col).value,2)

        for row in range(actual_row, actual_row + no_of_days):  # for 12M There are  365 days
            actual_interest = actual_sheet.cell(row=row, column=actual_col).value
            if expected_interest_rate != actual_interest:
                logger.error(f"Interest rate didn't match at row {row}: expected {expected_interest_rate}, got {actual_interest}")
                cls.error_msgs.append(f"Interest rate didn't match at row {row}: expected {expected_interest_rate}, got {actual_interest}")
            else:
               cls.matched_rows = True

    @classmethod
    def calculate_and_check_the_daily_interest_amount(cls, actual_sheet, actual_row, actual_col, no_of_days):
        interest_rate = float(actual_sheet.cell(row=actual_row, column=actual_col-1).value)

        for row in range(actual_row, actual_row + no_of_days):  # for 12M there are 365 days
            closing_balance = float(actual_sheet.cell(row=row, column=actual_col - 2).value)
            daily_interest_rate = (closing_balance * (interest_rate / 100) / 365)  # Calculation for daily interest rate for 12M (365 days)#

            actual_interest_rate = float(actual_sheet.cell(row=row, column=actual_col).value)
            if abs(actual_interest_rate - daily_interest_rate) > 1e-5:
                logger.error(f"Daily interest rate didn't match at row {row}: expected {daily_interest_rate}, got {actual_interest_rate}")
                cls.error_msgs.append(f"Daily interest amount didn't match at row {row}: expected {daily_interest_rate}, got {actual_interest_rate}")
            else:
                cls.matched_rows = True

    @classmethod
    def calculate_and_check_the_daily_interest_amount_cumulative(cls, actual_sheet, actual_row, actual_col, no_of_days):
        daily_interest_amount_cumulative = 0

        for row in range(actual_row, actual_row + no_of_days):  # for 12M there are 365 days
            daily_interest_amount = float(actual_sheet.cell(row=row, column=actual_col - 1).value)
            actual_interest_amount_cumulative = float(actual_sheet.cell(row=row, column=actual_col).value)
            daily_interest_amount_cumulative += daily_interest_amount
            if abs(actual_interest_amount_cumulative - daily_interest_amount_cumulative) > 1e-5:
                logger.error(f"Interest amount cumulative didn't match at row {row}: expected {daily_interest_amount_cumulative}, got {actual_interest_amount_cumulative}")
                cls.error_msgs.append(f"Interest amount cumulative didn't match at row {row}: expected {daily_interest_amount_cumulative}, got {actual_interest_amount_cumulative}")
            else:
                cls.matched_rows = True

    @classmethod
    def calculate_and_check_the_opening_balance_post_interest(cls, actual_sheet, actual_row, actual_col, no_of_days):

        for row in range(actual_row+1, actual_row + no_of_days-1):
            opening_balance_pre_interest = float(actual_sheet.cell(row=row,column=actual_col - 7).value)  # Getting the value of Opening Balance pre Interest
            daily_interest_amount_cumulative = float(actual_sheet.cell(row=row-1, column=actual_col - 1).value) #Getting the daily interest amount cumulative

            actual_opening_balance_post_interest = float(actual_sheet.cell(row=row, column=actual_col).value)
            expected_opening_balance_post_interest = opening_balance_pre_interest + daily_interest_amount_cumulative

            if abs(actual_opening_balance_post_interest - round(expected_opening_balance_post_interest,3)) > 1e-1:
                logger.error(f"Opening Balance Post Interest didn't match at row {row}: expected {expected_opening_balance_post_interest}, got {actual_opening_balance_post_interest}")
                cls.error_msgs.append(f"Opening Balance Post Interest didn't match at row {row}: expected {expected_opening_balance_post_interest}, got {actual_opening_balance_post_interest}")
            else:
                cls.matched_rows = True

    @classmethod
    def calculate_and_check_the_closing_balance_post_interest(cls, actual_sheet, actual_row, actual_col, no_of_days):

        for row in range(actual_row, actual_row + no_of_days):
            closing_balance_pre_interest = float(actual_sheet.cell(row=row, column=actual_col - 5).value)  # Getting the value of Closing Balance pre Interest
            daily_interest_amount_cumulative = float(actual_sheet.cell(row=row,column=actual_col - 2).value)  # Getting the daily interest amount cumulative

            actual_closing_balance_post_interest = float(actual_sheet.cell(row=row, column=actual_col).value)
            expected_closing_balance_post_interest = closing_balance_pre_interest + daily_interest_amount_cumulative

            if abs(actual_closing_balance_post_interest - round(expected_closing_balance_post_interest,3)) > 1e-1:
                logger.error(f"Closing Balance Post Interest didn't match at row {row}: expected {expected_closing_balance_post_interest}, got {actual_closing_balance_post_interest}")
                cls.error_msgs.append(f"Closing Balance Post Interest didn't match at row {row}: expected {expected_closing_balance_post_interest}, got {actual_closing_balance_post_interest}")
            else:
                cls.matched_rows = True

    @classmethod
    def calculate_and_check_the_daily_interest_amount_compounding(cls, actual_sheet, actual_row, actual_col, no_of_days):

        interest_rate = float(actual_sheet.cell(row=actual_row, column=actual_col-5).value)  # Getting the interest_rate

        for row in range(actual_row+1, actual_row + no_of_days-1):
            closing_balance_post_interest = round(float(actual_sheet.cell(row=row, column=actual_col - 1).value),2)  # Getting the value of Closing Balance pre Interest
            expected_compound_interest_amount = round((closing_balance_post_interest * (interest_rate / 100))/365 , 8)

            actual_compound_interest_amount = float(actual_sheet.cell(row=row, column=actual_col).value)

            if abs(actual_compound_interest_amount - expected_compound_interest_amount) > 1e-0:
                logger.error(f"Interest amount Compounding didn't match at row {row}: expected {expected_compound_interest_amount}, got {actual_compound_interest_amount}")
                cls.error_msgs.append(f"Interest amount Compounding didn't match at row {row}: expected {expected_compound_interest_amount}, got {actual_compound_interest_amount}")
            else:
                cls.matched_rows = True

    @classmethod
    def calculate_and_check_the_daily_interest_amount_compounding_cumulative(cls, actual_sheet, actual_row, actual_col, no_of_days):

        daily_interest_amount_cumulative = 0

        for row in range(actual_row, actual_row + no_of_days):
            daily_interest_amount = float(actual_sheet.cell(row=row, column=actual_col - 1).value)
            actual_interest_amount_cumulative = float(actual_sheet.cell(row=row, column=actual_col).value)
            daily_interest_amount_cumulative += daily_interest_amount
            if abs(actual_interest_amount_cumulative - daily_interest_amount_cumulative) > 1e-0:
                logger.error(f"Interest amount Cumulative Compounding didn't match at row {row}: expected {daily_interest_amount_cumulative}, got {actual_interest_amount_cumulative}")
                cls.error_msgs.append(f"Interest amount Cumulative Compounding didn't match at row {row}: expected {daily_interest_amount_cumulative}, got {actual_interest_amount_cumulative}")
            else:
               cls.matched_rows = True

    @classmethod
    def calculate_and_check_the_opening_balance_post_interest_compounding(cls, actual_sheet, actual_row, actual_col, no_of_days):

        for row in range(actual_row + 1, actual_row + no_of_days):
            opening_balance_pre_interest = float(actual_sheet.cell(row=row , column=actual_col - 11).value)  # Getting the value of Opening Balance pre Interest
            daily_interest_amount_cumulative = float(actual_sheet.cell(row=row - 1, column=actual_col - 1).value)  # Getting the daily interest amount cumulative

            actual_opening_balance_post_interest = float(actual_sheet.cell(row=row, column=actual_col).value)
            expected_opening_balance_post_interest = opening_balance_pre_interest + daily_interest_amount_cumulative

            if abs(actual_opening_balance_post_interest - expected_opening_balance_post_interest) > 1e-0:
                logger.error(f"Opening Balance Post Interest Compounding didn't match at row {row}: expected {expected_opening_balance_post_interest}, got {actual_opening_balance_post_interest}")
                cls.error_msgs.append(f"Opening Balance Post Interest Compounding didn't match at row {row}: expected {expected_opening_balance_post_interest}, got {actual_opening_balance_post_interest}")
            else:
               cls.matched_rows = True

    @classmethod
    def calculate_and_check_the_closing_balance_post_interest_compounding(cls, actual_sheet, actual_row, actual_col, no_of_days):

        for row in range(actual_row, actual_row + no_of_days):
            closing_balance_pre_interest = float(actual_sheet.cell(row=row, column=actual_col - 9).value)  # Getting the value of Closing Balance pre Interest
            daily_interest_amount_cumulative = float(actual_sheet.cell(row=row, column=actual_col - 2).value)  # Getting the daily interest amount cumulative

            actual_closing_balance_post_interest = float(actual_sheet.cell(row=row, column=actual_col).value)
            expected_closing_balance_post_interest = closing_balance_pre_interest + daily_interest_amount_cumulative

            if abs(actual_closing_balance_post_interest - expected_closing_balance_post_interest) > 1e-0:
                logger.error(f"Closing Balance Post Interest Compounding didn't match at row {row}: expected {expected_closing_balance_post_interest}, got {actual_closing_balance_post_interest}")
                cls.error_msgs.append(f"Closing Balance Post Interest Compounding didn't match at row {row}: expected {expected_closing_balance_post_interest}, got {actual_closing_balance_post_interest}")
            else:
               cls.matched_rows = True

    def test_to_validate_calculation_of_Daily_Interest_Rate_FTD(self):
        global match_rows_count
        all_expected_files = os.listdir(self.expected_data_dir)
        match_rows_count = 0

        # Read and aggregate data from all expected files
        for file_name in all_expected_files:
            expectedExcel_file_path = os.path.join(self.expected_data_dir, file_name)
            expected_sheet = self.load_expected_excel(expectedExcel_file_path)

            actualExcel_file_path = os.path.join(self.actual_data_dir, "SAVING_DEV.CONF.DAILY_INTEREST_RATE_FIXED_TERM.xlsx")
            actual_sheet = self.load_actual_excel(actualExcel_file_path)

            covered_reference_number = []

            if expected_sheet is not None:
                for exp_row in range(4, expected_sheet.max_row + 1):
                    expected_row_value = self.get_row_values(expected_sheet, exp_row, 2)
                    for actual_row in range(2, actual_sheet.max_row + 1):
                        actual_row_value = self.get_row_values(actual_sheet, actual_row, 3)

                        # for 12M there are 365 days and for 01M there are 30 days
                        if expected_row_value == actual_row_value and "12M" in actual_row_value and expected_row_value not in covered_reference_number:
                            self.check_the_date(expected_sheet, exp_row, 9, actual_sheet, actual_row, 4,365)  # in expected_sheet Date column is 9 and in actual_sheet 4
                            self.check_the_day_sequence(actual_sheet, actual_row, 5,365)  # check the Day Sequence 1,2,...,364,365
                            self.check_the_opening_balance_pre_interest(expected_sheet, exp_row, 8, actual_sheet,actual_row, 6,365)  # check the opening balance pre interest
                            self.check_the_credit(expected_sheet, exp_row, 8, actual_sheet, actual_row, 7)
                            self.check_the_closing_balance_pre_interest(expected_sheet, exp_row, 8, actual_sheet,actual_row, 9,365)  # check the closing balance pre interest
                            self.check_the_interest_rate(expected_sheet, exp_row, 7, actual_sheet, actual_row, 10,365)  # check the interest rate
                            self.calculate_and_check_the_daily_interest_amount(actual_sheet, actual_row, 11,365)  # calculate the daily interest amount
                            self.calculate_and_check_the_daily_interest_amount_cumulative(actual_sheet, actual_row,12,365)  # calculate the daily interest amount cumulative
                            self.calculate_and_check_the_opening_balance_post_interest(actual_sheet, actual_row, 13,365)  # calculate the opening balance pre interest
                            self.calculate_and_check_the_closing_balance_post_interest(actual_sheet, actual_row, 14,365)  # calculate the closing balance pre interest
                            self.calculate_and_check_the_daily_interest_amount_compounding(actual_sheet, actual_row,15,365)  # calculate the daily interest amount compounding
                            self.calculate_and_check_the_daily_interest_amount_compounding_cumulative(actual_sheet,actual_row,16,365)  # calculate the daily interest amount compound cumulative
                            self.calculate_and_check_the_opening_balance_post_interest_compounding(actual_sheet,actual_row, 17,365)  # Calculate the opening balance post interest compounding
                            self.calculate_and_check_the_closing_balance_post_interest_compounding(actual_sheet,actual_row, 18,365)  # Calculate the closing balance post interest compounding
                            if self.matched_rows:
                                match_rows_count += 1
                            covered_reference_number.append(expected_row_value)
                            break
                        elif expected_row_value == actual_row_value and "01M" in actual_row_value and expected_row_value not in covered_reference_number:
                            self.check_the_date(expected_sheet, exp_row, 9, actual_sheet, actual_row, 4,30)  # check for next 30 days, in expected_sheet Date column is 9 and in actual_sheet 4
                            self.check_the_day_sequence(actual_sheet, actual_row, 5,30)  # check the Day Sequence 1,2,...28,29,30
                            self.check_the_opening_balance_pre_interest(expected_sheet, exp_row, 8, actual_sheet,actual_row, 6,30)  # check the opening balance pre interest
                            self.check_the_closing_balance_pre_interest(expected_sheet, exp_row, 8, actual_sheet,actual_row, 9,30)  # check the closing balance pre interest
                            self.check_the_interest_rate(expected_sheet, exp_row, 7, actual_sheet, actual_row, 10,30)  # check the interest rate
                            self.calculate_and_check_the_daily_interest_amount(actual_sheet, actual_row, 11,30)  # calculate the daily interest amount
                            self.calculate_and_check_the_daily_interest_amount_cumulative(actual_sheet, actual_row,12,30)  # calculate the daily interest amount cumulative
                            self.calculate_and_check_the_opening_balance_post_interest(actual_sheet, actual_row, 13,30)  # calculate the opening balance pre interest
                            self.calculate_and_check_the_closing_balance_post_interest(actual_sheet, actual_row, 14,30)  # calculate the closing balance pre interest
                            self.calculate_and_check_the_daily_interest_amount_compounding(actual_sheet, actual_row,15,30)  # calculate the daily interest amount compounding
                            self.calculate_and_check_the_daily_interest_amount_compounding_cumulative(actual_sheet,actual_row,16,30)  # calculate the daily interest amount compound cumulative
                            self.calculate_and_check_the_opening_balance_post_interest_compounding(actual_sheet,actual_row, 17,30)  # Calculate the opening balance post interest compounding
                            self.calculate_and_check_the_closing_balance_post_interest_compounding(actual_sheet,actual_row, 18,30)  # Calculate the closing balance post interest compounding
                            if self.matched_rows:
                                match_rows_count += 1
                            covered_reference_number.append(expected_row_value)
                            break

            else:
                logger.error(f"Error loading expected sheet: {expectedExcel_file_path}")

        # self.error_msgs.append(f"Total rows: {self.total_rows}")
        # self.error_msgs.append(f"Matched rows: {match_rows_count}")
        # self.error_msgs.append(f"Unmatched rows: {self.total_rows-match_rows_count}")

        if self.error_msgs:
            # Join all error messages into a single message
            error_msg = "<br/>".join(self.error_msgs)
            # Fail the test with the error message
            self.fail(error_msg)
        else:
            self.assertTrue(True)

    @classmethod
    def tearDownClass(cls):
        logger.info("Tearing down resources")
        cls.clean_up_downloaded_files(cls.expected_data_dir)
        cls.clean_up_downloaded_files(cls.actual_data_dir)

    @classmethod
    def clean_up_downloaded_files(cls, folder_path):
        logger.info(f"Cleaning up files in {folder_path}...")
        files = glob.glob(os.path.join(folder_path, "*"))
        for file in files:
            os.remove(file)
            logger.info(f"Deleted file: {file}")
        logger.info("Clean up completed.")


if __name__ == '__main__':
    unittest.main()
